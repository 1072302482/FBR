import matplotlib
# matplotlib.use('Agg')  # Avoid no display error, and must be placed before other matplotlib-related imports
matplotlib.use('TkAgg')
import numpy as np
from numpy.lib.index_tricks import c_
from PIL.Image import init
import matplotlib.pyplot as plt
import matplotlib.colors
from scipy.io import wavfile
import scipy.signal
import scipy.fft
#import pdb;pdb.set_trace()         #この文をデバックしたいところにぶち込んで実行
from scipy.io.wavfile import write
import os
from auxiva_iss import auxiva_iss
from iva_goto import *
from gc_auxiva_iss import *
from bss_eval_sources_cpu import *

from function_off import *
from make_room import *
import openpyxl
import librosa
import librosa.display
from test import *


if __name__ == '__main__':
    ## Setting  ここをいじったら、パラメータ変更
    ## 変える ##
    n_mic = 2  # number of microphone, 2 ~ 6
    # DOA = [45, 90, 110, 130, 150, 170]  # 45°は確定、残りを90~150の中で入力。個数はマイク数。DOA of interference sign

    ############################################################################################################################################################################################################################
    DOA = np.array(np.linspace(10,70, num=5))
    # DOA = [20]
    ############################################################################################################################################################################################################################

    num_angles = 720
    DOA_target_range = np.linspace(0, 120, num_angles)
    DOA_interf_range = np.linspace(-180, 0, num_angles)
    lamb = [100]  # weight of GC term  const=unit, null時に使用
    dataset = "c"  # A, A2, B, Cのいずれかを代入。　A:45,90-150 B:45,90-170 C:0-80,100-180
    rt = 0.2  # 0.1 or 0.2 or 0.3   reverberation_time[ms]
    DOAestalg = "SRP" # "SRP" or "NormMUSIC" or "MUSIC"

    ############################################################################################################################################################################################################################
    alg = "GCcsv2"  # "csv"or "GCcsv"or "FBR"or "GCcsv2" or "pureFBR"手法
    ############################################################################################################################################################################################################################
    
    
    Bfl = 314 #Frame length in a block
    ##変えない##
    n_src = 6  # number of src, 2 (or 3 or 4まだ対応していません)
    const = "double"  # "unit" or "null" or "double"　制約３種類
    lamb_unit = [1]  # const=double時に使用
    lamb_null = [0]  # const=double時に使用
    lamb_sc = 0  # weight of SC（スケール）term (GC-AuxIVA-ISS is not supported)　基本０でよい
    
    fs = 16000  # frequency rate [Hz]
    n_fft = 2048  # fft length [samples]
    n_hop = 1024
    n_iter = 50  # number of iteration 更新回数
    dist_src = 1  # distance from mic to src [m]
    
    list_ave_tgt_fSDR, list_ave_tgt_fSIR, list_ave_tgt_fSAR, list_aveinf_fSDR, list_aveinf_fSIR, list_aveinf_fSAR, list_avetime, list_order = [], [], [], [], [], [], [], []
    list_ave_tgt_sSDR, list_ave_tgt_sSIR, list_ave_tgt_sSAR = [], [], []
    list_lu, list_ln = [], []
    num_l = len(lamb)
    num_lu = len(lamb_unit)
    num_ln = len(lamb_null)
    srcfile_num = []
    if n_mic == 2:                  ## positions of microphones
        d = np.asarray([-0.01, 0.01])
    elif n_mic == 3:
        d = np.asarray([-0.01, 0.01, 0.03])
    elif n_mic == 4:
        d = np.asarray([-0.03, -0.01, 0.01, 0.03])
    elif n_mic == 5:
        d = np.asarray([-0.03, -0.01, 0.01, 0.03, 0.05])
    elif n_mic == 6:
        d = np.asarray([-0.05, -0.03, -0.01, 0.01, 0.03, 0.05])
    else:
        raise ValueError("Number of microphones should be 2~6.")
    if n_mic == 2:  ## positions of microphones
        d1 = np.asarray([[-0.01, 0.01], [0, 0]])
    elif n_mic == 3:
        d1 = np.asarray([[-0.01, 0.01, 0.03], [0, 0, 0]])
    elif n_mic == 4:
        d1 = np.asarray([[-0.03, -0.01, 0.01, 0.03], [0, 0, 0, 0]])
    elif n_mic == 5:
        d1 = np.asarray([[-0.03, -0.01, 0.01, 0.03, 0.05], [0, 0, 0, 0, 0]])
    elif n_mic == 6:
        d1 = np.asarray([[-0.05, -0.03, -0.01, 0.01, 0.03, 0.05], [0, 0, 0, 0, 0, 0]])
    else:
        raise ValueError("Number of microphones should be 2~6.")
    print(d1)
    out_dir = "output2/" + str(dataset) + "_RT(" + str(int(rt * 1000)) + "ms)/" + "proposed method_" + str(
        n_src) + "src_" + str(n_mic) + "mic/"
    summary_dir = out_dir + str(alg) + "_" + "_data summary" + "/"
    os.makedirs(summary_dir, exist_ok=True)
    DOA = np.array(DOA)
    str_DOA = np.array2string(DOA, max_line_width=np.inf)
    
    ###########################################################################################################
    # double constraint
    if const == "double":
        for l_u in lamb_unit:
            for l_n in lamb_null:
                write_path_mat = out_dir + "mat/" + str_DOA + "/" + alg + "_" + const + "_λunit(" + str(
                    l_u) + "_λnull(" + str(l_n) + ")_"
                # write_path_mat = out_dir + "mat/" + '1' + "/" + alg + "_" + const + "_λunit(" + str(
                #     l_u) + "_λnull(" + str(l_n) + ")_"
                os.makedirs((out_dir + "mat/" + str_DOA + "/"), exist_ok=True)
                # os.makedirs((out_dir + "mat/" + '1' + "/"), exist_ok=True)
                list_tgt_fSDR, list_tgt_fSIR, list_tgt_fSAR, list_inf_fSDR, list_inf_fSIR, list_inf_fSAR, list_time = [], [], [], [], [], [], []
                list_sSDR, list_sSIR, list_sSAR, list_tgt_sSDR, list_tgt_sSIR, list_tgt_sSAR = [], [], [], [], [], []
                sSDR_all, sSIR_all, sSAR_all = [], [], []
                sSDRerr_all, sSIRerr_all, sSARerr_all = [], [], []
                cond = 1
                order_true = 0
         
                for snum in range(50):
                    # path setting
                    # data_dir = "data/" + str(dataset) + "/" + "RT(" + str(int(rt*1000)) + "ms)/" + str(n_src) + "src" + "6mic/"  + str(cond).zfill(3)"/"
                    data_dir = "data/" + str(dataset) + "/" + "2src_6mic_RT(" + str(int(rt * 1000)) + "ms)/" + str(
                        cond).zfill(3) + "/"
                    save_dir = out_dir + str(cond).zfill(3) + "/"
                    save_dir_bss = save_dir + alg + " " + const + "/" + "λ_u=" + str(l_u) + ", λ_n=" + str(
                        l_n) + ", DOA" + str_DOA + "/"
                    os.makedirs(data_dir, exist_ok=True)
                    os.makedirs(save_dir_bss, exist_ok=True)
                    
                    # read wavfile
                    _, mix = wavfile.read((data_dir + "mix.wav"))
                    mix = mix.T
                    mix = mix[:n_mic, :]
                    
                    _, ref_sig = wavfile.read((data_dir + "ref_sig.wav"))
                    ref_sig = ref_sig[:, np.newaxis]

                    n_ref_sig = ref_sig.shape[-1]

                    
                    # STFT
                    f, _, X = scipy.signal.stft(mix, fs=fs, nperseg=n_fft, noverlap=n_fft - n_hop)#(314,1025,6)(L,F,M)
                    X = X.transpose(2, 1, 0)[:, :, :n_mic]  # (frame, freq, mic)#(314,1025,6)(L,F,M)
                    X1 = X
                    # X = np.conj(X)
                    

                    array_DOAb_target = DOA
                    n_sv = len(array_DOAb_target)
                    sv = np.zeros((f.size, n_mic, n_sv), dtype="complex")
                    for i in range(n_sv):
                        sv[:, :, i, None] = generate_sv(d, array_DOAb_target[i], f2omega(f))  # (freq, mic, n_sv, None)
                        print(sv[f.size-1, :, i, None])
                        print('/')
                    # FBR
                    ###############################################################################################################################
                    if  'FBR' in alg:

                        h, h_SC = beamformtest2(X, 40, 140, d, 40, f2omega(f), f)# h (L,F,M)
                        # Y (F L N)
                        ## MVDR test
                        # h, hw = MVDR(X, sv, dic={})  # h (F x nSv x T)
                        # h = h.transpose(2, 0, 1)

                        X1 = h
                        # output the istft signal after each mic
                        h = h.transpose(1,0,2) # h (F,L,M)
                        # for i in range(h.shape[2]):
                        #     sigbf = scipy.signal.istft(h[:, :, i], fs, nperseg=n_fft, noverlap=n_fft - n_hop)[1]  #
                        #     # istftしている
                        #     # sig_write = sigbf / np.max(np.abs(sigbf)) * .9
                        #     sigbf = ((sigbf / np.max(np.abs(sigbf)) * 2 ** 15) * 0.9).astype(
                        #         dtype=ref_sig.dtype)  # クリッピング　（クリッピング、過大入力で検索）　あと量子化もしている2^15のところ
                        #     wavfile.write((save_dir_bss + alg + "_" + str(i) + ".wav"), fs, sigbf)
                       
                    ###############################################################################################################################

                    
                    
                    X1 = X1.transpose(1, 2, 0)
                    # Make block for CSV
                    XBlock = makeXBlock(X1, Bfl)  # (F x Nb x M x Bfl)
                    XBlock = XBlock.transpose(2, 0, 3, 1)#(M x F x Bfl x Nb)
                    M1, F1, BFl1, Nb1 = XBlock.shape
                    XBlock1 = np.conj(XBlock)
                    array_DOAb_target = estimate_DOA(d=d1, fs=fs, n_fft=n_fft, XBlock=XBlock1,
                                                     DOA_target=DOA_target_range,
                                                     DOAestalg="MUSIC",plot=False)
                    array_DOAb_interf = estimate_DOA(d=d1, fs=fs, n_fft=n_fft, XBlock=XBlock1,
                                                     DOA_target=DOA_interf_range,
                                                     DOAestalg="MUSIC",plot=False)
                    # array_DOAb_target = DOA
                    # n_sv = len(array_DOAb_target)
                    # sv = np.zeros((f.size, n_mic, n_sv), dtype="complex")
                    # for i in range(n_sv):
                    #     sv[:, :, i, None] = generate_sv(d, array_DOAb_target[i], f2omega(f))  # (freq, mic, n_sv, None)

                    Y_list = []
                    U_list = []
                    for Nb in range(Nb1):
                        # svb=sv[:, :, Nb, None]
                        svb = sv
                        # h_SC = beamform(svb, array_DOAb_target[Nb], DOA_interf_range[Nb], f)
                        #h_SC (freq, mic, mic)
                    # #
                    # plt.close()
                    # # h_SC = beamform(sv,40,140,f)
                    # # h_SC = np.array(h_SC)
                    
                    # setting lambda matrix
                        c = np.eye(svb.shape[2])
                        lamb_matrix = (l_u * np.eye(svb.shape[2])) + (l_n * (1 - np.eye(svb.shape[2])))
                        
                        # perform separation
                        
                        if alg == "csv":    # U  (offlineIter x F x M x N)
                            Y, U, time_list = CSV_AuxIVE(XBlock[:, :, :, Nb])
                            
                        elif alg == "GCcsv":
                            Y, U, time_list = gc_CSV_AuxIVE(XBlock[:,:,:,Nb], sv=svb,
                                                            save_dir=save_dir_bss)
                        elif alg == "GCcsv2":
                            Y, U, time_list = GCCSVAuxIVE(XBlock[:,:,:,Nb], sv=svb)
                        elif alg == "FBR":
                            
                            # Y, U, time_list = CSV_AuxIVE(XBlock[:, :, :, Nb])
                            Y, U, time_list = GCCSVAuxIVE(XBlock[:, :, :, Nb], sv=svb )
                            # Y = Y
                            #Y (F N L)
                        if 'pure' not in alg:
                            Y_list.append(Y)
                            U_list.append(U) #Wb(100, 1025, 6,1)
                            list_time.append(np.mean(time_list))
                    if 'pure' not in alg:
                        # Convert lists to arrays and stack along a new dimension
                        Y_array = np.stack(Y_list, axis=2)  # New shape: (Nb, ...)
                        U_array = np.stack(U_list, axis=4)  # New shape: (Nb, ...)
                        Y_array_reshaped = Y_array.reshape(Y_array.shape[0], Y_array.shape[1], -1)
                        U_array_reshaped = U_array #Ub(50, 1025, 6,1,3)
                        Y = Y_array_reshaped
                        U = U_array_reshaped
                        Y = Y.transpose(0,2,1)# Y (F L N)
                        W = U[-1, :, :, :, -1]  # W  (F x M x N)
                        W = W.transpose(0, 2, 1)
                        # W = np.conj(W)
                        # W (F N M)
                    if alg == "pureFBR":
                        Y = np.sum(h, axis=2)
                        Y = Y[ : ,:, np.newaxis]
                        W = h_SC[ :, np.newaxis ,:]
                        # W = np.conj(W)
                    for i in range(Y.shape[2]):
                        sig = scipy.signal.istft(Y[:, :, i], fs, nperseg=n_fft, noverlap=n_fft - n_hop)[1]  # istftしている
                        # sig_write = sig / np.max(np.abs(sig)) * .9
                        sig = ((sig / np.max(np.abs(sig)) * 2 ** 15) * 0.9).astype(
                            dtype=ref_sig.dtype)  # クリッピング　（クリッピング、過大入力で検索）　あと量子化もしている2^15のところ
                        wavfile.write((save_dir_bss + alg + "_" + str(i) + ".wav"), fs, sig)
                        est_sig = sig[:len(ref_sig), None] if i == 0 else np.concatenate(
                            (est_sig, sig[:len(ref_sig), None]), axis=1)

                        # plot signals and directivity patterns for investigation
                        pattern, f_idx = directivity(W[:, i, None, :], f, d)[0:2]
                        # plot_signal(sig, (save_dir_bss + "sig_" + str(i) + ".png"))
                        if alg == 'csv':
                            write('output_no_gc.wav', fs, sig)
                            plot_signal(sig, path='S_no_gc.png')
                            plot_directivity(pattern[:, :, 0], vmin=-10, vmax=12, cmap='seismic',
                                             path='D_no_gc_parttern.png')
                        if alg == 'GCcsv':
                            write('output_gc.wav', fs, sig)
                            plot_signal(sig, path='S_gc.png')
                            plot_directivity(pattern[:, :, 0], vmin=-10, vmax=12, cmap='seismic',
                                             path='D_gc_parttern.png')
                        if alg == 'GCcsv2':
                            write('output_gc2.wav', fs, sig)
                            plot_signal(sig, path='S_gc2.png')
                            plot_directivity(pattern[:, :, 0], vmin=-20, vmax=-0, cmap='seismic',
                                             path='D_gc2_parttern.png')
                        if alg == 'FBR':
                            write('output_FBR.wav', fs, sig)
                            plot_signal(sig, path='S_FBR.png')
                            plot_directivity(pattern[:, :, 0], vmin=-20, vmax= 20, cmap='seismic',
                                             path='D_FBR_parttern.png')
                        if alg == 'pureFBR':
                            write('output_FBR_pure.wav', fs, sig)
                            plot_signal(sig, path='S_FBR_pure.png')
                            plot_directivity(pattern[:, :, 0], vmin=-10, vmax= 10, cmap='seismic',
                                             path='D_FBR_parttern_pure.png')
                        # plot_directivity(pattern[:, :, 0], vmin=-10, vmax=12, cmap='seismic', path='D_parttern.png')
                        # plot_directivity(pattern[:, :, 0], path=(save_dir_bss + "directivity_" + str(i) + ".png"))
                       
                        # plot_directivity(pattern[:, :, 0], vmin=-10, vmax=12, cmap='seismic', path=None)
                        # plt.savefig('D_parttern.png')
                        
                    
                    # evaluation
                    # full SDR
                    L_est_sig = est_sig.shape[0]
                    SDR_array, SIR_array, SAR_array, popt = bss_eval_sources(est_sig.T, ref_sig[:L_est_sig,0].T)
                    # order_true = order_true + 1 if (popt == np.arange(n_src)).all() else order_true
                    tgt_fSDR = SDR_array[0]
                    tgt_fSIR = SIR_array[0]
                    tgt_fSAR = SAR_array[0]
       
                    print(" tgt_fSDR",tgt_fSDR," tgt_fSIR",tgt_fSIR," tgt_fSAR",tgt_fSAR,"popt",popt)
                    # segmental SDR
                    SDREachTime, SIREachTime, SAREachTime,  popts = segement_bss_eval_sources_online(
                        est_sig.T, ref_sig.T, fs, flen=512)
                    sSDR, sSIR, sSAR = SDREachTime.mean(), SIREachTime.mean(), SAREachTime.mean()
                    sSDR_array, sSIR_array, sSAR_array = SDREachTime.mean(0), SIREachTime.mean(0), SAREachTime.mean(0)
                    tgt_sSDR = sSDR_array[0]
                    tgt_sSIR = sSIR_array[0]
                    tgt_sSAR = sSAR_array[0]
                    print(" tgt_sSDR", tgt_sSDR, " tgt_sSIR", tgt_sSIR, " tgt_sSAR", tgt_sSAR, "popt", popts)
                    list_sSDR.append(SDR_array)
                    list_sSIR.append(SIR_array)
                    list_sSAR.append(SAR_array)
                    sSDR_all.append(SDREachTime)
                    sSIR_all.append(SIREachTime)
                    sSAR_all.append(SAREachTime)

                    
                    ## write score
                    list_tgt_fSDR.append(tgt_fSDR)
                    list_tgt_fSIR.append(tgt_fSIR)
                    list_tgt_fSAR.append(tgt_fSAR)
  
                    
                    list_tgt_sSDR.append(tgt_sSDR)
                    list_tgt_sSIR.append(tgt_sSIR)
                    list_tgt_sSAR.append(tgt_sSAR)
                    
                    with open((save_dir_bss + "score.txt"), mode='w') as file:
                        file.write(
                            f"SDR: {np.float16(SDR_array)}\nSIR: {np.float16(SIR_array)}\nSAR: {np.float16(SAR_array)}\niteration: {n_iter}\n")
                        file.write(
                            f"constraint: {const}\nλ(unit): {l_u}, λ(null): {l_n}\nλSC: {lamb_sc}\nDOA: {DOA}\norder: {popt}")
                    
                    write_path = out_dir + "/" + alg + "_" + const + "_λ(unit)[" + str(l_u) + "]_λ(null)[" + str(
                        l_n) + "]_"
                    
                    # order writing
                    with open((write_path + "order_true.txt"), mode='w') as file:
                        file.write(f"{order_true}")
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    
                    # score writing
                    for i in range(len(list_tgt_fSDR)):
                        sheet.cell(row=i + 1, column=1).value = list_tgt_fSDR[i]
                        sheet.cell(row=i + 1, column=2).value = list_tgt_fSIR[i]
                        sheet.cell(row=i + 1, column=3).value = list_tgt_fSAR[i]
                    wb.save(write_path + "score.xlsx")
                    
                    # time writing
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    sheet.title = alg
                    for i in range(len(list_time)):
                        sheet.cell(row=i + 1, column=1).value = list_time[i]
                    wb.save(write_path + "time.xlsx")
                    
                    print(f"λ_u={l_u}, λ_n={l_n}: {cond}/50 is completed.")
                    
                    cond += 1
                
                # summary data
                ave_tgt_fSDR = np.mean(list_tgt_fSDR)
                ave_tgt_fSAR = np.mean(list_tgt_fSAR)
                ave_tgt_fSIR = np.mean(list_tgt_fSIR)
                aveinf_fSDR = np.mean(list_inf_fSDR)
                aveinf_fSAR = np.mean(list_inf_fSAR)
                aveinf_fSIR = np.mean(list_inf_fSIR)
                avetime = np.mean(list_time)
                ave_sSDR = np.mean(list_sSDR,axis=0)
                ave_sSIR = np.mean(list_sSIR,axis=0)
                ave_sSAR = np.mean(list_sSAR,axis=0)
                ave_tgt_sSDR = np.mean(list_tgt_sSDR)
                ave_tgt_sSIR = np.mean(list_tgt_sSIR)
                ave_tgt_sSAR = np.mean(list_tgt_sSAR)
                
                list_ave_tgt_fSDR.append(ave_tgt_fSDR)
                list_ave_tgt_fSIR.append(ave_tgt_fSIR)
                list_ave_tgt_fSAR.append(ave_tgt_fSAR)
                list_aveinf_fSDR.append(aveinf_fSDR)
                list_aveinf_fSIR.append(aveinf_fSIR)
                list_aveinf_fSAR.append(aveinf_fSAR)
                list_avetime.append(avetime)
                list_order.append(order_true)
                list_ave_tgt_sSDR.append(ave_tgt_sSDR)
                list_ave_tgt_sSIR.append(ave_tgt_sSIR)
                list_ave_tgt_sSAR.append(ave_tgt_sSAR)
                
                # mat
                sSDR_all = np.array(sSDR_all).mean(0)  # SDR_all ax0:cond, ax1:time, ax2=ch ??
                sSIR_all = np.array(sSIR_all).mean(0)
                sSAR_all = np.array(sSAR_all).mean(0)
                sSDRerr_all = np.array(sSDRerr_all).mean(0)
                sSIRerr_all = np.array(sSIRerr_all).mean(0)
                sSARerr_all = np.array(sSARerr_all).mean(0)
                # for s in range(n_src):
                #     ws_mat = write_path_mat + str(s) + "_"
                #     bssEvalOnline(sSDR_all[:, s], sSDRerr_all, sSIR_all[:, s], sSIRerr_all, sSAR_all[:, s], sSARerr_all,
                #                   ws_mat)
        
        ###  summary.xlsx  ###
        wb = openpyxl.Workbook()
        ws = wb.active
        # setting height & weight
        ws.row_dimensions[1].height = 20
        for i, c in enumerate('abcdefghij', start=1):
            ws.column_dimensions[c].width = 10
        
        ## index
        ws['A1'] = 'λ_unit'
        ws['B1'] = 'λ_null'
        ws['C1'] = 'sSDR'
        ws['D1'] = 'sSIR'
        ws['E1'] = 'sSAR'
        ws['F1'] = 'fullSDR'
        ws['G1'] = 'fullSIR'
        ws['H1'] = 'fullSAR'
        ws['I1'] = 'Ave_Time'
        ws['J1'] = 'order'
        
        # input dat
        for i in range(1, num_l + 1):
            ws.cell(i + 1, column=1).value = list_lu[i - 1]
            ws.cell(i + 1, column=2).value = list_ln[i - 1]
            ws.cell(i + 1, column=3).value = list_ave_tgt_sSDR[i - 1]  # sSDR (tgt)
            ws.cell(i + 1, column=4).value = list_ave_tgt_sSIR[i - 1]  # sSIR (tgt)
            ws.cell(i + 1, column=5).value = list_ave_tgt_sSAR[i - 1]  # sSAR (tgt)
            ws.cell(i + 1, column=6).value = list_ave_tgt_fSDR[i - 1]  # fullSDR (tgt)
            ws.cell(i + 1, column=7).value = list_ave_tgt_fSIR[i - 1]  # fullSIR (tgt)
            ws.cell(i + 1, column=8).value = list_ave_tgt_fSAR[i - 1]  # fullSAR (tgt)
            ws.cell(i + 1, column=9).value = list_avetime[i - 1]  # ave_time
            ws.cell(i + 1, column=10).value = list_order[i - 1]  # order
        wb.save(summary_dir + f"{const}_DOA{DOA}.xlsx")
    
    
    
    
  