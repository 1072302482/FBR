import scipy.signal
import scipy.fft
#import pdb;pdb.set_trace()         #この文をデバックしたいところにぶち込んで実行
from scipy.io.wavfile import write
import os
import pandas as pd

from bss_eval_sources_cpu import *

from function_off import *

import openpyxl

from test import *


if __name__ == '__main__':
    # 指定 RT 和 alg 的候选值
    RT_values = [0.1, 0.2, 0.3]
    alg_values = ["csv", "GCcsv", "pureFBR"]
    
# for rt in RT_values:
# for alg in alg_values:
    ## Setting  ここをいじったら、パラメータ変更
    ## 変える ##
    n_mic = 6  # number of microphone, 2 ~ 6
    # DOA = [45, 90, 110, 130, 150, 170]  # 45°は確定、残りを90~150の中で入力。個数はマイク数。DOA of interference sign

    ############################################################################################################################################################################################################################
    DOA = np.array(np.linspace(10,70, num=5))
    # DOA = [20]
    ############################################################################################################################################################################################################################

    num_angles = 720
    DOA_target_range = np.linspace(0, 120, num_angles)
    DOA_interf_range = np.linspace(-180, 0, num_angles)
    lamb = [100]  # weight of GC term  const=unit, null時に使用
    dataset = "c"  # A, A2, B, Cのいずれかを代入。　A:45,90-150 B:45,90-170 C:0-80,100-180
    rt = 0.3  # 0.1 or 0.2 or 0.3   reverberation_time[ms]
    DOAestalg = "SRP" # "SRP" or "NormMUSIC" or "MUSIC"

    ############################################################################################################################################################################################################################
    alg = "GCFBRcsv"  # "csv"or "OCsv"or "GCFBRcsv"or "OGCCsv"or "OFBRcsv"or "GCcsv"or "WPEcsv" or "pureFBR" or "GCWPEcsv" or
    # "GCFBRcsv"手法
    ###########################################################################################################################################################################################################################
    
    Bfl = 314 #Frame length in a block
    ##変えない##
    n_src = 6  # number of src, 2 (or 3 or 4まだ対応していません)
    const = "double"  # "unit" or "null" or "double"　制約３種類
    lamb_unit = [1]  # const=double時に使用
    lamb_null = [0]  # const=double時に使用
    lamb_sc = 0  # weight of SC（スケール）term (GC-AuxIVA-ISS is not supported)　基本０でよい
    
    fs = 16000  # frequency rate [Hz]
    n_fft = 2048  # fft length [samples]
    n_hop = 1024
    n_iter = 50  # number of iteration 更新回数
    # dist_src = 1  # distance from mic to src [m]
    
    list_ave_tgt_fSDR, list_ave_tgt_fSIR, list_ave_tgt_fSAR, list_aveinf_fSDR, list_aveinf_fSIR, list_aveinf_fSAR, list_avetime, list_order = [], [], [], [], [], [], [], []
    list_ave_tgt_sSDR, list_ave_tgt_sSIR, list_ave_tgt_sSAR = [], [], []
    list_lu, list_ln = [], []
    num_l = len(lamb)
    num_lu = len(lamb_unit)
    num_ln = len(lamb_null)
    srcfile_num = []
    if n_mic == 2:                  ## positions of microphones
        d = np.asarray([-0.01, 0.01])
    elif n_mic == 3:
        d = np.asarray([-0.01, 0.01, 0.03])
    elif n_mic == 4:
        d = np.asarray([-0.03, -0.01, 0.01, 0.03])
    elif n_mic == 5:
        d = np.asarray([-0.03, -0.01, 0.01, 0.03, 0.05])
    elif n_mic == 6:
        d = np.asarray([-0.05, -0.03, -0.01, 0.01, 0.03, 0.05])
    else:
        raise ValueError("Number of microphones should be 2~6.")
    if n_mic == 2:  ## positions of microphones
        d1 = np.asarray([[-0.01, 0.01], [0, 0]])
    elif n_mic == 3:
        d1 = np.asarray([[-0.01, 0.01, 0.03], [0, 0, 0]])
    elif n_mic == 4:
        d1 = np.asarray([[-0.03, -0.01, 0.01, 0.03], [0, 0, 0, 0]])
    elif n_mic == 5:
        d1 = np.asarray([[-0.03, -0.01, 0.01, 0.03, 0.05], [0, 0, 0, 0, 0]])
    elif n_mic == 6:
        d1 = np.asarray([[-0.05, -0.03, -0.01, 0.01, 0.03, 0.05], [0, 0, 0, 0, 0, 0]])
    else:
        raise ValueError("Number of microphones should be 2~6.")
    print(d1)
    print(alg)
    # out_dir = "evaluation_of_" + alg + "/"+str(dataset) + "_RT(" + str(int(rt * 1000)) + "ms)/" + \
    #           "proposed method_"\
    #           + \
    #           str(n_src) + "src_" + str(n_mic) + "mic/"
    out_dir = "evaluation_of_" + alg + "/"+str(dataset) + "_RT(" + str(int(rt * 1000)) + "ms)/" + \
              str(n_src) + "src_" + str(n_mic) + "mic/"
    summary_dir = out_dir + str(alg) + "_" + "_data summary" + "/"
    os.makedirs(summary_dir, exist_ok=True)
    DOA = np.array(DOA)
    str_DOA = np.array2string(DOA, max_line_width=np.inf)
    # 创建具有多级索引的空 DataFrame
    df = pd.DataFrame()
    
    ###########################################################################################################
    # double constraint
    if const == "double":
        for l_u in lamb_unit:
            for l_n in lamb_null:
                write_path_mat = out_dir + "mat/" + str_DOA + "/" + alg + "_" + const + "_λunit(" + str(
                    l_u) + "_λnull(" + str(l_n) + ")_"
                # write_path_mat = out_dir + "mat/" + '1' + "/" + alg + "_" + const + "_λunit(" + str(
                #     l_u) + "_λnull(" + str(l_n) + ")_"
                os.makedirs((out_dir + "mat/" + str_DOA + "/"), exist_ok=True)
                # os.makedirs((out_dir + "mat/" + '1' + "/"), exist_ok=True)
                list_tgt_fSDR, list_tgt_fSIR, list_tgt_fSAR, list_inf_fSDR, list_inf_fSIR, list_inf_fSAR, list_time = [], [], [], [], [], [], []
                list_sSDR, list_sSIR, list_sSAR, list_tgt_sSDR, list_tgt_sSIR, list_tgt_sSAR = [], [], [], [], [], []
                list_tgt_popt = []
                list_tgt_seg_popt = []
                sSDR_all, sSIR_all, sSAR_all = [], [], []
                # sSDRerr_all, sSIRerr_all, sSARerr_all = [], [], []
                ###############################################################################################################################################################################
                cond = 1
                order_true = 0
                df = pd.DataFrame()
                for snum in  range(2): #range(50):
                    # path setting
                    # data_dir = "data/" + str(dataset) + "/" + "RT(" + str(int(rt*1000)) + "ms)/" + str(n_src) + "src" + "6mic/"  + str(cond).zfill(3)"/"
                    data_dir = "data/" + str(dataset) + "/" + "2src_6mic_RT(" + str(int(rt * 1000)) + "ms)/" + str(
                        cond).zfill(3) + "/"
                    save_dir = out_dir + str(cond).zfill(3) + "/"
                    save_dir_bss = save_dir + alg + " " + const + "/" + "λ_u=" + str(l_u) + ", λ_n=" + str(
                        l_n) + ", DOA" + str(len(DOA)) + "/"
                    os.makedirs(data_dir, exist_ok=True)
                    os.makedirs(save_dir_bss, exist_ok=True)
                    
                    # read wavfile
                    _, mix = wavfile.read((data_dir + "mix.wav"))
                    mix = mix.T
                    mix = mix[:n_mic, :]
                    
                    _, ref_sig = wavfile.read((data_dir + "ref_sig.wav"))
                    ref_sig = ref_sig[:, np.newaxis]

                    n_ref_sig = ref_sig.shape[-1]

                    
                    # STFT
                    f, _, X = scipy.signal.stft(mix, fs=fs, nperseg=n_fft, noverlap=n_fft - n_hop)#(M,F,L)
                    X = X.transpose(2, 1, 0)[:, :, :n_mic]  # (frame, freq, mic)#(314,1025,6)(L,F,M)
                    X1 = X
                    # X = np.conj(X)
                    
                    # Use DOA to get steering vector
                    array_DOAb_target = DOA
                    n_sv = len(array_DOAb_target)
                    sv = np.zeros((f.size, n_mic, n_sv), dtype="complex")
                    for i in range(n_sv):
                        sv[:, :, i, None] = generate_sv(d, array_DOAb_target[i], f2omega(f))  # (freq, mic, n_sv, None)

                    # FBR
                    ###############################################################################################################################
                    if  'FBR' in alg:
                        h, h_SC, list_time,Gamma_target,Gamma_interfer = beamformtest1(X, 140, 40, d, 40,
                                                                                       f2omega(f),f)  # h (L,F,M)
                    ###############################################################################################################################

                    X1 = X1.transpose(1, 2, 0)
                    # Make block for CSV
                    XBlock = makeXBlock(X1, Bfl)  # (F x Nb x M x Bfl)
                    XBlock = XBlock.transpose(2, 0, 3, 1)#(M x F x Bfl x Nb)
                    M1, F1, BFl1, Nb1 = XBlock.shape
                    # DOA estimation Block by block
                    # array_DOAb_target = estimate_DOA(d=d1, fs=fs, n_fft=n_fft, XBlock=XBlock,
                    #                                  DOA_range=DOA_target_range,
                    #                                  DOAestalg=DOAestalg, plot=True)
                    # array_DOAb_interf = estimate_DOA(d=d1, fs=fs, n_fft=n_fft, XBlock=XBlock,
                    #                                  DOA_range=DOA_interf_range,
                    #                                  DOAestalg=DOAestalg, plot=True)
                    # array_DOAb_target = DOA
                    # n_sv = len(array_DOAb_target)
                    # sv = np.zeros((f.size, n_mic, n_sv), dtype="complex")
                    # for i in range(n_sv):
                    #     sv[:, :, i, None] = generate_sv(d, array_DOAb_target[i], f2omega(f))  # (freq, mic, n_sv, None)

                    Y_list = []
                    U_list = []
                    #Apply FBRGCCSV on each block
                    for Nb in range(Nb1):
                        # svb=sv[:, :, Nb, None]
                        svb = sv

                    # setting lambda matrix
                        c = np.eye(svb.shape[2])
                        lamb_matrix = (l_u * np.eye(svb.shape[2])) + (l_n * (1 - np.eye(svb.shape[2])))
                        
                        # perform separation
                        
                        if alg == "csv":    # U  (offlineIter x F x M x N)
                            Y, U, time_list = CSV_AuxIVE(XBlock[:, :, :, Nb])
                        elif alg == "OCsv":
                            Y, U, time_list = OCSVAuxIVE(X = XBlock[:, :, :, Nb])
                            U = U[np.newaxis, ...]
                        elif alg == "OGCCsv":
                            Y, U, time_list = OGCCSVAuxIVE(X = XBlock[:, :, :, Nb], sv=svb)
                            U = U[np.newaxis, ...]
                        elif alg == "OFBRcsv":
                            Y, U, time_list = OFBRCSVAuxIVE(X = XBlock[:,:,:,Nb], sv=svb, Gamma_target =
                            Gamma_target, Gamma_interfer = Gamma_interfer)
                            U = U[np.newaxis, ...]
                        elif alg == "GCcsv":
                            Y, U, time_list = GCCSVAuxIVE(XBlock[:,:,:,Nb], sv=svb)
                        elif alg == "WPEcsv":
                            Y, U, time_list = CSVWPEIVE(X = XBlock[:,:,:,Nb])
                        elif alg == "GCWPEcsv":
                            Y, U, time_list = GCCWPECSVIVE(X = XBlock[:,:,:,Nb], sv=svb)
                        elif alg == "GCFBRcsv":
                            # Y, U, time_list = FBRGCCSVAuxIVE2(X = XBlock[:,:,:,Nb], sv=svb, Gamma_target =
                            # Gamma_target, Gamma_interfer = Gamma_interfer)
                            Y, U, time_list = FBRGCCSVAuxIVE3(X = XBlock[:,:,:,Nb],filterFBR = h_SC )
                            # Y = Y
                            #Y (F N L)
                        if 'pure' not in alg:
                            Y_list.append(Y)
                            U_list.append(U) #Wb(100, 1025, 6,1)
                            list_time.append(np.sum(time_list))
                    if 'pure' not in alg:
                        # Convert lists to arrays and stack along a new dimension
                        Y_array = np.stack(Y_list, axis=2)  # New shape: (Nb, ...)
                        U_array = np.stack(U_list, axis=4)  # New shape: (Nb, ...)
                        Y_array_reshaped = Y_array.reshape(Y_array.shape[0], Y_array.shape[1], -1)
                        U_array_reshaped = U_array #Ub(50, 1025, 6,1,3)
                        Y = Y_array_reshaped
                        U = U_array_reshaped
                        Y = Y.transpose(0,2,1)# Y (F L N)
                        W = U[-1, :, :, :, -1]  # W  (F x M x N)
                        W = W.transpose(0, 2, 1)
                        # W = np.conj(W)
                        # W (F N M)
                    # Only process the input signal with FBR beamformer
                    if alg == "pureFBR":
                        h = h.transpose(1, 0, 2)  # h (F,L,M)
                        Y = np.sum(h, axis=2)
                        Y = Y[ : ,:, np.newaxis]
                        W = h_SC[ :, np.newaxis ,:]
                        # W = np.conj(W)
                    
                    for i in range(Y.shape[2]):
                        sig = scipy.signal.istft(Y[:, :, i], fs, nperseg=n_fft, noverlap=n_fft - n_hop)[1]  # istftしている
                        # sig_write = sig / np.max(np.abs(sig)) * .9
                        sig = ((sig / np.max(np.abs(sig)) * 2 ** 15) * 0.9).astype(
                            dtype=ref_sig.dtype)  # クリッピング　（クリッピング、過大入力で検索）　あと量子化もしている2^15のところ
                        wavfile.write((save_dir_bss + alg + "_" + str(i) + ".wav"), fs, sig)
                        est_sig = sig[:len(ref_sig), None] if i == 0 else np.concatenate(
                            (est_sig, sig[:len(ref_sig), None]), axis=1)

                        # plot signals and directivity patterns for investigation
                        pattern, f_idx = directivity(W[:, i, None, :], f, d)[0:2]
                        # plot_signal(sig, (save_dir_bss + "sig_" + str(i) + ".png"))
                        if alg == 'csv':
                            write('output_no_gc.wav', fs, sig)
                            plot_signal(sig, path='S_no_gc.png')
                            plot_directivity(pattern[:, :, 0], vmin=-10, vmax=12, cmap='seismic',
                                             path='D_no_gc_parttern.png')
                        if alg == 'GCcsv':
                            write('output_gc.wav', fs, sig)
                            plot_signal(sig, path='S_gc.png')
                            plot_directivity(pattern[:, :, 0], vmin=-20, vmax=0, cmap='seismic',
                                             path='D_gc_parttern.png')
                        if alg in ['OGCCsv', 'OFBRcsv']:
                            write('output_gc_OL.wav', fs, sig)
                            plot_signal(sig, path='S_gc_OL.png')
                            plot_directivity(pattern[:, :, 0], vmin=0, vmax=20, cmap='seismic',
                                             path='D_gc_parttern_OL.png')

                        if alg == 'pureFBR':
                            write('output_FBR_pure.wav', fs, sig)
                            plot_signal(sig, path='S_FBR_pure.png')
                            plot_directivity(pattern[:, :, 0], vmin=-20, vmax= 10, cmap='seismic',
                                             path='D_FBR_parttern_pure.png')
                        elif 'FBR' in alg:
                            write('output_FBR.wav', fs, sig)
                            plot_signal(sig, path='S_FBR.png')
                            plot_directivity(pattern[:, :, 0], vmin=-20, vmax= 5, cmap='seismic',
                                             path='D_FBR_parttern.png')
                        
                        
                        # plot_directivity(pattern[:, :, 0], vmin=-10, vmax=12, cmap='seismic', path='D_parttern.png')
                        # plot_directivity(pattern[:, :, 0], path=(save_dir_bss + "directivity_" + str(i) + ".png"))
                       
                        # plot_directivity(pattern[:, :, 0], vmin=-10, vmax=12, cmap='seismic', path=None)
                        # plt.savefig('D_parttern.png')Z
                        
                    
                    # evaluation
                    # full SDR
                    L_est_sig = est_sig.shape[0]
                    SDR_array, SIR_array, SAR_array, popt = bss_eval_sources(est_sig.T, ref_sig[:L_est_sig,0].T)
                    popt = popt.mean()
                    tgt_fSDR = SDR_array[0]
                    tgt_fSIR = SIR_array[0]
                    tgt_fSAR = SAR_array[0]
       
                    print(" tgt_fSDR",tgt_fSDR," tgt_fSIR",tgt_fSIR," tgt_fSAR",tgt_fSAR,"popt",popt)
                    # segmental SDR
                    SDREachTime, SIREachTime, SAREachTime,  popts = segement_bss_eval_sources_online(
                        est_sig.T, ref_sig.T, fs, flen=512)
                    seg_popts = popts
                    popts = popts.mean()
                    sSDR, sSIR, sSAR = SDREachTime.mean(), SIREachTime.mean(), SAREachTime.mean()
                    sSDR_array, sSIR_array, sSAR_array = SDREachTime.mean(0), SIREachTime.mean(0), SAREachTime.mean(0)
                    tgt_sSDR = sSDR_array[0]
                    tgt_sSIR = sSIR_array[0]
                    tgt_sSAR = sSAR_array[0]
                    print(" tgt_sSDR", tgt_sSDR, " tgt_sSIR", tgt_sSIR, " tgt_sSAR", tgt_sSAR, "popt", popts)
                    list_sSDR.append(SDR_array)
                    list_sSIR.append(SIR_array)
                    list_sSAR.append(SAR_array)
                    list_tgt_popt.append(popts)
                    sSDR_all = np.append(sSDR_all, SDREachTime)
                    sSIR_all = np.append(sSIR_all, SIREachTime)
                    sSAR_all = np.append(sSAR_all, SAREachTime)
                    
                    # 将数组转为 1-D，并创建 DataFrame
                    data = pd.DataFrame({'SDR': SDREachTime.flatten(),
                                         'SIR': SIREachTime.flatten(),
                                         'SAR': SAREachTime.flatten(),
                                         'PERM': seg_popts.flatten()})
                    
                    # 将 DataFrame 转置，使其有20列
                    data = data.T
                    
                    # 将新数据追加到旧 DataFrame
                    df = df.append(data)

                    ## write score
                    list_tgt_fSDR.append(tgt_fSDR)
                    list_tgt_fSIR.append(tgt_fSIR)
                    list_tgt_fSAR.append(tgt_fSAR)
    
                    
                    list_tgt_sSDR.append(tgt_sSDR)
                    list_tgt_sSIR.append(tgt_sSIR)
                    list_tgt_sSAR.append(tgt_sSAR)
                    list_tgt_popt.append(popts)
                    print(f"λ_u={l_u}, λ_n={l_n}: {cond}/50 is completed.")
                    cond += 1
                    
                df.to_excel(summary_dir + "_" + alg+"_" + f"{const}_DOA{len(DOA)}_1s_Online_evaluation.xlsx")
                # 加载工作簿并选择工作表

                
                with open((save_dir_bss + "score.txt"), mode='w') as file:
                    file.write(
                        f"SDR: {np.float16(SDR_array)}\nSIR: {np.float16(SIR_array)}\nSAR: {np.float16(SAR_array)}\niteration: {n_iter}\n")
                    file.write(
                        f"constraint: {const}\nλ(unit): {l_u}, λ(null): {l_n}\nλSC: {lamb_sc}\nDOA: {DOA}\norder: {popt}")
                
                write_path = out_dir + "/" + alg + "_" + const + "_λ(unit)[" + str(l_u) + "]_λ(null)[" + str(
                    l_n) + "]_"
                
                # order writing
                with open((write_path + "order_true.txt"), mode='w') as file:
                    file.write(f"{order_true}")
                wb = openpyxl.Workbook()
                wc = wb.active
                sheet = wb.active
                sheet.title = alg
                wc['A1'] = 'fullSDR'
                wc['B1'] = 'fullSIR'
                wc['C1'] = 'fullSAR'
                wc['D1'] = 'order'
                wc['E1'] = 'time'
                # score writing

                for i in (1, num_l + 1):
                    sheet.cell(row=i + 1, column=1).value = list_tgt_fSDR[i-1]
                    sheet.cell(row=i + 1, column=2).value = list_tgt_fSIR[i-1]
                    sheet.cell(row=i + 1, column=3).value = list_tgt_fSAR[i-1]
                    sheet.cell(row=i + 1, column=4).value = list_tgt_popt[i-1]
                wb.save(write_path + "score.xlsx")
                
                list_tgt_sSIR = [x for x in list_tgt_sSIR if x <= 100]
                list_tgt_fSIR = [x for x in list_tgt_fSIR if x <= 100]
                

                
                # summary data
                ave_tgt_fSDR = np.mean(list_tgt_fSDR)
                ave_tgt_fSAR = np.mean(list_tgt_fSAR)
                ave_tgt_fSIR = np.mean(list_tgt_fSIR)
                ave_tgt_fSIR = np.mean(list_tgt_fSIR)
                avetime = np.sum(list_time)
                ave_sSDR = np.mean(list_sSDR,axis=0)
                ave_sSIR = np.mean(list_sSIR,axis=0)
                ave_sSAR = np.mean(list_sSAR,axis=0)
                ave_tgt_sSDR = np.mean(list_tgt_sSDR)
                ave_tgt_sSIR = np.mean(list_tgt_sSIR)
                ave_tgt_sSAR = np.mean(list_tgt_sSAR)
                ave_tgt_popt = np.mean(list_tgt_popt)

                list_ave_tgt_fSDR.append(ave_tgt_fSDR)
                print("Appending data")
                list_ave_tgt_fSIR.append(ave_tgt_fSIR)
                list_ave_tgt_fSAR.append(ave_tgt_fSAR)

                list_avetime.append(avetime)
                list_order.append(order_true)
                list_ave_tgt_sSDR.append(ave_tgt_sSDR)
                list_ave_tgt_sSIR.append(ave_tgt_sSIR)
                list_ave_tgt_sSAR.append(ave_tgt_sSAR)
                
                # mat
                sSDR_all = np.array(sSDR_all).mean(0)  # SDR_all ax0:cond, ax1:time, ax2=ch
                sSIR_all = np.array(sSIR_all).mean(0)
                sSAR_all = np.array(sSAR_all).mean(0)

                # for s in range(n_src):
                #     ws_mat = write_path_mat + str(s) + "_"
                #     bssEvalOnline(sSDR_all[:, s], sSDRerr_all, sSIR_all[:, s], sSIRerr_all, sSAR_all[:, s], sSARerr_all,
                #                   ws_mat)
            
        ###  summary.xlsx  ###
        wb = openpyxl.Workbook()
        ws = wb.active
        # setting height & weight
        ws.row_dimensions[1].height = 20
        for i, c in enumerate('abcdefghij', start=1):
            ws.column_dimensions[c].width = 10
        
        ## index
        ws['A1'] = 'λ_unit'
        ws['B1'] = 'λ_null'
        ws['C1'] = 'sSDR'
        ws['D1'] = 'sSIR'
        ws['E1'] = 'sSAR'
        ws['F1'] = 'fullSDR'
        ws['G1'] = 'fullSIR'
        ws['H1'] = 'fullSAR'
        ws['I1'] = 'Ave_Time'
        ws['J1'] = 'order'
        
        # input dat
        for i in range(1, num_l + 1):
            ws.cell(i + 1, column=1).value = lamb_unit[0]
            ws.cell(i + 1, column=2).value = lamb_null[0]
            ws.cell(i + 1, column=3).value = list_ave_tgt_sSDR[i - 1]  # sSDR (tgt)
            ws.cell(i + 1, column=4).value = list_ave_tgt_sSIR[i - 1]  # sSIR (tgt)
            ws.cell(i + 1, column=5).value = list_ave_tgt_sSAR[i - 1]  # sSAR (tgt)
            ws.cell(i + 1, column=6).value = list_ave_tgt_fSDR[i - 1]  # fullSDR (tgt)
            ws.cell(i + 1, column=7).value = list_ave_tgt_fSIR[i - 1]  # fullSIR (tgt)
            ws.cell(i + 1, column=8).value = list_ave_tgt_fSAR[i - 1]  # fullSAR (tgt)
            ws.cell(i + 1, column=9).value = ave_tgt_popt  # order
            ws.cell(i + 1, column=10).value = list_avetime[0]  # ave_time

        wb.save(summary_dir + f"{const}_DOA{len(DOA)}.xlsx")
        
    
    
    
  